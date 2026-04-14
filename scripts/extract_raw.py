#!/usr/bin/env python3
"""
Extrahiert alle Rohdaten aus den FAKT II-Quellen (PDFs + Excel).
Gibt strukturierte Daten als JSON auf stdout aus, die von build_wiki.py konsumiert werden.

Verwendung:
    uv run --with pdfplumber --with openpyxl python3 scripts/extract_raw.py > /tmp/fakt_extracted.json

Quellen:
    raw/fakt_broschuere_1.pdf  – Allgemeine Infos, Antragstellung
    raw/fakt_broschuere_2.pdf  – Übersichtstabelle aller Maßnahmen
    raw/fakt_broschuere_3.pdf  – Detailbeschreibungen aller Maßnahmen
    raw/Kombinationstabelle FAKT II.xlsx – Kombinationsmatrix
"""

import json
import os
import sys

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
RAW = os.path.join(BASE, "raw")


def extract_pdfs():
    """Extrahiert Text aus allen drei PDFs."""
    import pdfplumber

    result = {}
    for name in ["fakt_broschuere_1.pdf", "fakt_broschuere_2.pdf", "fakt_broschuere_3.pdf"]:
        path = os.path.join(RAW, name)
        if not os.path.exists(path):
            print(f"WARNUNG: {path} nicht gefunden", file=sys.stderr)
            continue
        pages = []
        with pdfplumber.open(path) as pdf:
            for i, page in enumerate(pdf.pages):
                text = page.extract_text() or ""
                pages.append({"seite": i + 1, "text": text})
        result[name] = pages
    return result


def extract_kombinationstabelle():
    """Extrahiert die Kombinationsmatrix aus der Excel-Datei."""
    import openpyxl

    path = os.path.join(RAW, "Kombinationstabelle FAKT II.xlsx")
    if not os.path.exists(path):
        print(f"WARNUNG: {path} nicht gefunden", file=sys.stderr)
        return {}

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    # Header: Spalte D onwards = Maßnahmen-Codes (Zeile 1)
    header_codes = {}
    for col in range(4, ws.max_column + 1):  # D=4
        val = ws.cell(row=1, column=col).value
        if val:
            header_codes[col] = val.strip()

    # Beschreibungen (Zeile 3)
    header_titles = {}
    for col in range(4, ws.max_column + 1):
        val = ws.cell(row=3, column=col).value
        if val:
            header_titles[col] = val.strip().replace("\n", " ")

    # Fördersätze (Zeile 4)
    header_rates = {}
    for col in range(4, ws.max_column + 1):
        val = ws.cell(row=4, column=col).value
        if val:
            header_rates[col] = str(val).strip()

    # Matrix-Daten: Zeilen 5-46
    matrix = {}
    for row in range(5, ws.max_row + 1):
        row_code = ws.cell(row=row, column=1).value
        row_title = ws.cell(row=row, column=2).value
        if not row_code:
            continue
        row_code = str(row_code).strip()
        row_title = str(row_title).strip().replace("\n", " ") if row_title else ""

        combos = {}
        for col, col_code in header_codes.items():
            val = ws.cell(row=row, column=col).value
            if val is not None:
                combos[col_code] = str(val).strip()

        matrix[row_code] = {
            "titel": row_title,
            "kombinationen": combos,
        }

    # Fußnoten (Zeilen 48-54)
    fussnoten = []
    for row in range(48, 55):
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=row, column=col).value
            if val and len(str(val).strip()) > 10:
                fussnoten.append(str(val).strip().replace("\n", " "))

    return {
        "header_codes": header_codes,
        "header_titles": header_titles,
        "header_rates": header_rates,
        "matrix": matrix,
        "fussnoten": fussnoten,
    }


def main():
    data = {
        "pdfs": extract_pdfs(),
        "kombinationstabelle": extract_kombinationstabelle(),
    }
    json.dump(data, sys.stdout, ensure_ascii=False, indent=2)


if __name__ == "__main__":
    main()
